"""Readers for the ILS ceramide ring trial source tables.

All file I/O for the dataset generator lives here. Every reader returns plain
dicts keyed by natural composite keys, so the entry builders stay pure.

LabId is the join key everywhere -- including the original lab report workbooks,
whose filenames carry it. The xlsx writes unpadded numerics ('3') while the CSVs
write zero-padded ones ('03'), so every LabId is pushed through
normalize_lab_id() on read.
"""

import csv
import re

from openpyxl import load_workbook

import ringtrial_terms as terms

LAB_ID_RE = re.compile(r"^(\d+)([ab]?)$")
MISSING = {"", "NA", "ND", "NaN", "None"}

# Results table 3: lab x material x ceramide (extraction replicates averaged).
T3_LAB_ID, T3_SAMPLE_TYPE, T3_CERAMIDE = 1, 2, 3
T3_SP_MEAN, T3_SP_SD, T3_MP_MEAN, T3_MP_SD = 12, 13, 14, 15

# Results table 2: + one row per extraction replicate (injections averaged).
T2_LAB_ID, T2_SAMPLE_TYPE, T2_SAMPLE, T2_CERAMIDE = 1, 2, 3, 4
T2_SP_MEAN, T2_MP_MEAN = 13, 15

_XLSX_TO_CODE = {spec["xlsx"]: code for code, spec in terms.MATERIALS.items()}


def normalize_lab_id(raw):
    """'3' -> '03', '10a' -> '10a', '02b' -> '02b'."""
    text = str(raw).strip()
    match = LAB_ID_RE.match(text)
    if not match:
        raise ValueError(f"unrecognised LabId: {raw!r}")
    return f"{int(match.group(1)):02d}{match.group(2)}"


def _material_code(sample_type):
    """'NIST SRM' -> 'SRM'. Raises on an unknown reference material."""
    code = _XLSX_TO_CODE.get(str(sample_type).strip())
    if code is None:
        raise ValueError(f"unknown SampleType: {sample_type!r}")
    return code


def _optional(value):
    """Return a stripped string, or None if the cell is a missing-value marker."""
    if value is None:
        return None
    text = str(value).strip()
    return None if text in MISSING else text


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_lab_metadata(definitions_dir):
    """Join instruments.csv onto LabProtocolDetails_Manuscript.csv by LabId.

    LabProtocolDetails_Manuscript.csv defines the 39 laboratories in the
    manuscript; instruments.csv also lists labs excluded from it, so the
    protocol table drives the join.
    """
    instruments = {}
    with open(definitions_dir / "instruments.csv", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            instruments[normalize_lab_id(row["LabId"])] = row

    metadata = {}
    with open(definitions_dir / "LabProtocolDetails_Manuscript.csv",
              encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            lab_id = normalize_lab_id(row["LabId"])
            instrument = instruments.get(lab_id)
            if instrument is None:
                raise ValueError(f"LabId {lab_id} missing from instruments.csv")
            metadata[lab_id] = {
                "labNum": row["LabNum"].strip(),
                "protocol": row["Protocol"].strip(),
                "lc": row["LC"].strip(),
                "resolution": row["MassAnalyzerResolution"].strip(),
                "gradientTime": _optional(row["GradientTime"]),
                "sourceTemp": _optional(row["SourceTemp"]),
                "vendor": instrument["Vendor"].strip(),
                "instrument": instrument["Instrument"].strip(),
                "analyzer": instrument["MassAnalyzerType"].strip(),
            }
    return metadata


def _rows(xlsx_path, sheet_name):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        yield from list(workbook[sheet_name].iter_rows(values_only=True))[1:]
    finally:
        workbook.close()


def read_lab_stats(xlsx_path):
    """Per-laboratory mean and SD over extraction replicates (Results table 3)."""
    stats = {}
    for row in _rows(xlsx_path, "Results table 3"):
        key = (normalize_lab_id(row[T3_LAB_ID]),
               _material_code(row[T3_SAMPLE_TYPE]),
               str(row[T3_CERAMIDE]).strip())
        stats[key] = {
            "single-point": {"mean": _number(row[T3_SP_MEAN]),
                             "sd": _number(row[T3_SP_SD])},
            "multi-point": {"mean": _number(row[T3_MP_MEAN]),
                            "sd": _number(row[T3_MP_SD])},
        }
    return stats


def read_replicates(xlsx_path):
    """Per-extraction-replicate concentrations (Results table 2)."""
    replicates = {}
    for row in _rows(xlsx_path, "Results table 2"):
        key = (normalize_lab_id(row[T2_LAB_ID]),
               _material_code(row[T2_SAMPLE_TYPE]),
               str(row[T2_CERAMIDE]).strip())
        replicates.setdefault(key, []).append({
            "sample": str(row[T2_SAMPLE]).strip(),
            "single-point": _number(row[T2_SP_MEAN]),
            "multi-point": _number(row[T2_MP_MEAN]),
        })
    for entries in replicates.values():
        entries.sort(key=lambda entry: entry["sample"])
    return replicates


def read_outliers(output_dir):
    """Outlier flags per calibration method (Suppl_Dataset_* CSVs)."""
    files = {
        "multi-point": ("Suppl_Dataset_Concentrations_OutlierStatus_"
                        "Sheet-MultiPointConc.csv"),
        "single-point": ("Suppl_Dataset_Concentrations_OutlierStatus_"
                         "Sheet-SinglePointConc.csv"),
    }
    flags = {}
    for method, filename in files.items():
        with open(output_dir / filename, encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                key = (normalize_lab_id(row["LabId"]),
                       row["SampleType"].strip(),
                       row["CeramideName"].strip(),
                       method)
                flags[key] = row["Outlier"].strip().upper() == "TRUE"
    return flags


def read_consensus(output_dir):
    """Published across-laboratory consensus values (Suppl_TableS01).

    Verified multi-point: the All and Filt means reproduce the mean of the
    multi-point per-lab values to sixteen decimal places.
    """
    path = (output_dir / "Suppl_TableS01_Concentrations-CVs_"
                         "ALL-vs-Outlierfilt-datasets.csv")
    fields = ("n", "mean", "sd", "median", "cv", "rcv")
    consensus = {}
    with open(path, encoding="utf-8-sig") as handle:
        rows = [row for row in csv.reader(handle) if row and any(row)]
    for row in rows[1:]:
        material, ceramide = row[0].strip(), row[1].strip()
        for variant, offset in (("All", 2), ("Filt", 8)):
            values = dict(zip(fields, (float(cell) for cell in row[offset:offset + 6])))
            values["n"] = int(values["n"])
            consensus[(material, ceramide, variant)] = values
    return consensus
